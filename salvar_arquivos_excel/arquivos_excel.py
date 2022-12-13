from openpyxl import Workbook, workbook
from openpyxl import load_workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


analise_entradas = 'Analise_Ope_Entrada'
analise_saidas = 'Analise_Ope_Saida'
analise_full = 'Analise_Ope_Entrada_Saida'
numeracao = 'Numeracao_NFe_NFCe'
Numeracao_CTe = 'Numeracao_CTe'
sped_sefaz_e = 'Sped_x_Sefaz_Entrada'
sefaz_sped_e = 'Sefaz_x_Sped_Entrada'
sped_sefaz_cte_e = 'Sped_x_Sefaz_CTe_Entrada'
sefaz_sped_cte_e = 'Sefaz_x_Sped_CTe_Entrada'
sped_sefaz_s = 'Sped_x_Sefaz_Saida'
sefaz_sped_s = 'Sefaz_x_Sped_Saida'
sped_sefaz_cte_s = 'Sped_x_Sefaz_CTe_Saida'
sefaz_sped_cte_s = 'Sefaz_x_Sped_CTe_Saida'
Fecha_Saidas = 'Fecha_Saidas'
compara_entradas = 'Diferenca_Entradas'
compara_saidas = 'Diferenca_Saidas'
compara_saidas_cte = 'Diferenca_CTe_Saidas'
compara_entradas_cte = 'Diferenca_CTe_Entrada'
C100_x_C190_Saidas = 'C100_x_C190_Saidas'
C100_x_C190_Entradas = 'C100_x_C190_Entradas'
Fecha_CTe_Saida = 'Fecha_CTe_Saida'


def formata_numero(ws, intervalo):
    columns = ws[intervalo]

    for rows in columns:
        for cell in rows:
            cell.number_format = '#,##0.00'


def autosize(ws):
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True


def teste_sheet(ws, nome):
    if ws.title != 'Sheet':
        return True
    else:
        ws.title = nome
        return False


def salvar_analise_Full(lista_entradas, lista_resumo):
    wb = load_workbook('power_fisco.xlsx')

    ws = wb.active

    if teste_sheet(ws, analise_full):
        ws = wb.create_sheet(analise_full)

    i = 1
    for l in lista_entradas:
        ws.append(l)

    ws[f'T{i}'] = 'CFOP'
    ws[f'U{i}'] = 'CST'
    ws[f'V{i}'] = 'ALIQ'
    ws[f'W{i}'] = 'VLR CONTÁBIL'
    ws[f'X{i}'] = 'BC ICMS'
    ws[f'Y{i}'] = 'ICMS'
    ws[f'Z{i}'] = 'E115'
    i += 1

    for l in lista_resumo:
        ws[f'T{i}'] = l[0]
        ws[f'U{i}'] = l[1]
        ws[f'V{i}'] = l[2]
        ws[f'W{i}'] = l[3]
        ws[f'X{i}'] = l[4]
        ws[f'Y{i}'] = l[5]
        ws[f'Z{i}'] = l[6]
        i += 1

    formata_numero(ws, 'W:Z')
    autosize(ws)
    ws.auto_filter.ref = 'A:Z'
    wprops = ws.sheet_properties
    wprops.tabColor = "FFFFFF00"
    wb.save('power_fisco.xlsx')


def salvar_analise_c170(lista_entradas, lista_resumo):
    wb = load_workbook('power_fisco.xlsx')

    ws = wb.active

    if teste_sheet(ws, analise_entradas):
        ws = wb.create_sheet(analise_entradas)

    i = 1
    for l in lista_entradas:
        ws.append(l)

    ws[f'T{i}'] = 'CFOP'
    ws[f'U{i}'] = 'CST'
    ws[f'V{i}'] = 'ALIQ'
    ws[f'W{i}'] = 'VLR CONTÁBIL'
    ws[f'X{i}'] = 'BC ICMS'
    ws[f'Y{i}'] = 'ICMS'
    ws[f'Z{i}'] = 'E115'

    i += 1

    for l in lista_resumo:
        ws[f'T{i}'] = l[0]
        ws[f'U{i}'] = l[1]
        ws[f'V{i}'] = l[2]
        ws[f'W{i}'] = l[3]
        ws[f'X{i}'] = l[4]
        ws[f'Y{i}'] = l[5]
        ws[f'Z{i}'] = l[6]
        i += 1

    formata_numero(ws, 'W:Z')
    autosize(ws)
    ws.auto_filter.ref = 'A:Z'
    wprops = ws.sheet_properties
    wprops.tabColor = "FFFFFF00"
    wb.save('power_fisco.xlsx')


def salvar_analise_c190(lista_entradas, lista_resumo):
    wb = load_workbook('power_fisco.xlsx')

    ws = wb.active

    if teste_sheet(ws, analise_saidas):
        ws = wb.create_sheet(analise_saidas)

    i = 1
    for l in lista_entradas:
        ws.append(l)

    ws[f'T{i}'] = 'CFOP'
    ws[f'U{i}'] = 'CST'
    ws[f'V{i}'] = 'ALIQ'
    ws[f'W{i}'] = 'VLR CONTÁBIL'
    ws[f'X{i}'] = 'BC ICMS'
    ws[f'Y{i}'] = 'ICMS'
    ws[f'Z{i}'] = 'E115'

    i += 1

    for l in lista_resumo:
        ws[f'T{i}'] = l[0]
        ws[f'U{i}'] = l[1]
        ws[f'V{i}'] = l[2]
        ws[f'W{i}'] = l[3]
        ws[f'X{i}'] = l[4]
        ws[f'Y{i}'] = l[5]
        ws[f'Z{i}'] = l[6]
        i += 1

    formata_numero(ws, 'W:Z')
    autosize(ws)

    ws.auto_filter.ref = 'A:Z'
    wprops = ws.sheet_properties
    wprops.tabColor = "FFFFFF00"
    wb.save('power_fisco.xlsx')


def salvar_pulou_nf(lista_entradas, CTE=False):
    wb = load_workbook('power_fisco.xlsx')

    ws = wb.active

    name_arquivo = Numeracao_CTe if CTE else numeracao

    if teste_sheet(ws, name_arquivo):
        ws = wb.create_sheet(name_arquivo)

    if not lista_entradas:
        ws["a1"] = 'Numero_NF'
        ws["b1"] = 'Modelo'
        ws["c1"] = 'Série'
        wprops = ws.sheet_properties
        wprops.tabColor = "3CB371"
        ws["B5"] = "Não há documentos fiscais faltantes nos registros do SPED!"
        ws["B5"].font = Font(bold=True)
        ws["B5"].font = Font(size='20')
    else:
        i = 2
        ws["a1"] = 'Numero_NF'
        ws["b1"] = 'Modelo'
        ws["c1"] = 'Série'
        for l in lista_entradas:
            # ws['a' + str(i)] = l
            ws.append(l)
            i += 1
        wprops = ws.sheet_properties
        wprops.tabColor = "FF0000"
    ws.auto_filter.ref = 'A:C'

    wb.save('power_fisco.xlsx')


def salvar_spedxsefaz(lista_entradas, eous, cte=False):

    if eous == '0':
        if cte:
            nome = sped_sefaz_cte_e
        else:
            nome = sped_sefaz_e
    else:
        if cte:
            nome = sped_sefaz_cte_s
        else:
            nome = sped_sefaz_s

    wb = load_workbook('power_fisco.xlsx')

    ws = wb.active

    if teste_sheet(ws, nome):
        ws = wb.create_sheet(nome)

    for l in lista_entradas:
        ws.append(l)

    ws.auto_filter.ref = 'A:K'

    wprops = ws.sheet_properties

    if len(lista_entradas) == 1:
        wprops.tabColor = "3CB371"
        ws["B5"] = "Todas NFe/NFCe do SPED encontram-se no SEFAZ!" if not cte else "Todos CTe do SPED encontram-se no SEFAZ"
        ws["B5"].font = Font(bold=True)
        ws["B5"].font = Font(size='20')
    else:
        wprops.tabColor = "FF0000"

    wb.save('power_fisco.xlsx')


def salvar_energia_eletrica(lista_entradas):

    nome = 'Energia_Eletrica'

    wb = load_workbook('power_fisco.xlsx')

    ws = wb.active

    if teste_sheet(ws, nome):
        ws = wb.create_sheet(nome)

    for li in lista_entradas:
        ws.append(*li)

    # ws.auto_filter.ref = 'A:K'

    wprops = ws.sheet_properties

    if len(lista_entradas) == 1:
        wprops.tabColor = "FF0000"
        ws["B5"] = "Não há nenhuma nota de Energia Elétrica Lançada!"
        ws["B5"].font = Font(bold=True)
        ws["B5"].font = Font(size='20')
    else:

        wprops.tabColor = "3CB371"

    wb.save('power_fisco.xlsx')


def salvar_sefazxsped(lista_entradas, eous, cte=False):

    if eous == '0':
        if cte:
            nome = sefaz_sped_cte_e
        else:
            nome = sefaz_sped_e
    else:
        if cte:
            nome = sefaz_sped_cte_s
        else:
            nome = sefaz_sped_s

    wb = load_workbook('power_fisco.xlsx')

    ws = wb.active

    if teste_sheet(ws, nome):
        ws = wb.create_sheet(nome)

    for l in lista_entradas:
        ws.append(l)

    ws.auto_filter.ref = 'A:U'
    wprops = ws.sheet_properties

    if len(lista_entradas) == 1:
        wprops.tabColor = "3CB371"
        ws["B5"] = "Todas NFe/NFCe do SEFAZ encontram-se no SPED!" if not cte else "Todos CTe do SEFAZ encontram-se no SPED"
        ws["B5"].font = Font(bold=True)
        ws["B5"].font = Font(size='20')
    else:
        wprops.tabColor = "FF0000"

    wb.save('power_fisco.xlsx')


def salvar_fechamento_saidas(lista_entradas, msg, cte=False):
    wb = load_workbook('power_fisco.xlsx')

    ws = wb.active

    if cte:
        nome = Fecha_CTe_Saida
    else:
        nome = Fecha_Saidas

    if teste_sheet(ws, nome):
        ws = wb.create_sheet(nome)

    for l in lista_entradas:
        ws.append(l)

    # print(ws['b:f'])
    columns = ws['b:f']

    for rows in columns:
        for cell in rows:
            cell.number_format = '#,##0.00'

    # print(ws.max_column)
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True
        # ws.column_dimensions[get_column_letter(i)].auto_size = True

    wprops = ws.sheet_properties

    if not msg:
        wprops.tabColor = "3CB371"
    else:
        wprops.tabColor = "FF0000"

    wb.save('power_fisco.xlsx')


def salvar_compara_entradas(lista_entradas, eous, cte=False):
    wb = load_workbook('power_fisco.xlsx')

    ws = wb.active

    if eous == '0':
        if cte:
            nome = compara_entradas_cte
        else:
            nome = compara_entradas
    else:
        if cte:
            nome = compara_saidas_cte
        else:
            nome = compara_saidas

    if teste_sheet(ws, nome):
        ws = wb.create_sheet(nome)

    for l in lista_entradas:
        ws.append(l)

    wprops = ws.sheet_properties
    ws.auto_filter.ref = 'A:W'

    if len(lista_entradas) == 1:
        wprops.tabColor = "3CB371"
        ws["B5"] = "Não há diferença de Valor, Base de Cáculo, ICMS e ICMS-ST entre SEFAZ e SPED!" if not cte else "Não há diferença de Valor, Base de Cáculo e ICMS entre SEFAZ e SPED!"
        ws["B5"].font = Font(bold=True)
        ws["B5"].font = Font(size='20')
    else:
        wprops.tabColor = "FF0000"

    wb.save('power_fisco.xlsx')


def salvar_compara_c100xc190(lista_entradas, eous):
    wb = load_workbook('power_fisco.xlsx')

    ws = wb.active

    if eous == '1':
        compara = C100_x_C190_Saidas
    else:
        compara = C100_x_C190_Entradas

    if teste_sheet(ws, compara):
        ws = wb.create_sheet(compara)

    for l in lista_entradas:
        ws.append(l)

    wprops = ws.sheet_properties
    ws.auto_filter.ref = 'A:Q'

    if len(lista_entradas) == 1:
        wprops.tabColor = "3CB371"
        ws["B5"] = "Não há diferença de Valor, Base de Cáculo, ICMS e ICMS-ST entre SEFAZ e SPED!"
        ws["B5"].font = Font(bold=True)
        ws["B5"].font = Font(size='20')
    else:
        wprops.tabColor = "FF0000"

    wb.save('power_fisco.xlsx')
